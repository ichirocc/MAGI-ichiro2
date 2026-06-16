#!/usr/bin/env python3
# Updated MAGI UI/UX diff report — re-assessed against branch claude/code-review-0gu5p3 @ e4f6de8.
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

wb = openpyxl.Workbook()
HEAD = "a38fcbf"  # main へマージ済み（WorkManager + ログ強化 まで反映）
BRANCH = "claude/code-review-0gu5p3 → main"

hdr = Font(bold=True, color="FFFFFF")
hfill = PatternFill("solid", fgColor="0F6E66")
ok = PatternFill("solid", fgColor="CDEFE0")
part = PatternFill("solid", fgColor="F7E6BE")
rem = PatternFill("solid", fgColor="F8D6D2")
wrap = Alignment(wrap_text=True, vertical="top")
title = Font(bold=True, size=13)

def sheet(name, headers, rows, widths, status_col=None):
    ws = wb.create_sheet(name)
    for c, h in enumerate(headers, 1):
        cell = ws.cell(1, c, h); cell.font = hdr; cell.fill = hfill; cell.alignment = wrap
    for r, row in enumerate(rows, 2):
        for c, v in enumerate(row, 1):
            cell = ws.cell(r, c, v); cell.alignment = wrap
        if status_col:
            st = str(row[status_col-1])
            f = ok if st.startswith("解決") else part if st.startswith("部分") else rem if st.startswith("残") else None
            if f: ws.cell(r, status_col).fill = f
    for c, w in enumerate(widths, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(c)].width = w
    return ws

# 00 サマリ
ws = wb.active; ws.title = "00_サマリ(更新)"
ws["A1"] = "MAGI UI/UX 差異レポート（更新版）"; ws["A1"].font = title
summary = [
    ("更新日時", "2026-06-12"),
    ("比較対象(更新)", f"{BRANCH} @ {HEAD}（旧レポートは 511a1f7 基準）"),
    ("Web版", "magi_v6_web.html (v5.33.19)"),
    ("", ""),
    ("旧レポート差異総数", "121（主な差異43＋片方のみ20＋文言20＋挙動20＋データ18）"),
    ("本ブランチで 解決(✅)", "約16件（担当外警告・修正導線・重大のみ・触覚・残り時間・操作ログJSON・背景最適化 ほか）"),
    ("本ブランチで 部分対応(🟡)", "約6件（中断警告はBG継続+通知で改善・非色手がかり・自動保存文言 ほか）"),
    ("残存(❌)", "約8件（ws6/ws7・希望上書き・月次/年次wizard・他の案・フッター履歴 ほか）"),
    ("追加実装(仕様書)", "WorkManager背景最適化+完了通知 / 5仮説ハイブリッド早期キャンセル / R8 Keep / 操作監査ログ"),
    ("", ""),
    ("方針(継続)", "業務導線・長時間実行・違反表示はWebへ寄せる。Android固有(ファイル/テーマ/片手/触覚/ボトムシート)は維持し文言統一。"),
    ("最優先の残課題", "①WorkManager+完了通知(中断耐性) ②希望で上書き(確認+Undo+ログ) ③ws6/ws7分離 ④非色手がかり(実線/破線)"),
]
for i, (k, v) in enumerate(summary, 3):
    ws.cell(i, 1, k).font = Font(bold=True); ws.cell(i, 2, v).alignment = wrap
ws.column_dimensions["A"].width = 26; ws.column_dimensions["B"].width = 90

# 01 本ブランチ対応一覧
sheet("01_本ブランチ対応一覧",
      ["分類", "実装した操作/UX", "効果(差分の解消)", "関連コミット"],
      [
        ["スマホ操作", "セルtap→ボトムシート選択", "セル編集方式を選択式で確定(誤巡回防止)", "b95f827"],
        ["スマホ操作", "ハプティクス/スワイプ日付送り/片手モード", "指1本操作特化(Web比でネイティブ優位を維持)", "b95f827"],
        ["操作コパイロット", "満足度ゲージ/ライブ進捗(残り時間)", "進捗可視化でガチャ操作抑止(Web BusyOverlay相当)", "2b92e26/39f1d47"],
        ["操作コパイロット", "ガチャ助言/研磨限界の合図", "同設定再実行を検知し緩和候補を提示", "2b92e26"],
        ["Web操作移植", "担当外希望の警告+希望編集へ誘導", "片方のみ#4(希望サマリ/担当外警告)を解消", "39f1d47"],
        ["Web操作移植", "修正導線(編集タブへジャンプ)", "片方のみ#5(NextActionBar)を解消", "39f1d47"],
        ["Web操作移植", "違反内訳『重大のみ』フィルタ", "片方のみ#8/重大差異(重大度識別)を一部解消", "39f1d47"],
        ["エンジン", "最大5仮説並列+ハイブリッド早期キャンセル", "省電力(合格で残り即停止)・相対評価で時間内最良採用", "51f9964"],
        ["エンジン", "適応的ALNS(2023研究)", "オペレータ学習重みで探索効率向上", "f4c9341"],
        ["堅牢性", "R8 keep(model)", "難読化でスキーマ/JSON・CSVが壊れない(§8)", "51f9964"],
        ["校正", "タップ目標48dp統一/文言重複の解消", "指1本操作の合格基準・表記ゆれ解消", "e4f6de8"],
        ["品質保証", "Webゴールデンパリティテスト(CI)", "解決層一致・HARD=0配布判定一致を恒久ガード", "5f5b8e7/a23650a"],
        ["マルチタスク", "WorkManager背景最適化+完了通知", "アプリを閉じても計算継続(中断耐性向上)・完了通知(§6)", "1ae0228"],
        ["監査", "操作ログ(追記式)+JSON出力", "いつ何をして結果どうなったかを追える監査ログ(片方のみ#14解消)", "a38fcbf"],
      ],
      [14, 36, 46, 14])

# 02 重大差異の更新
sheet("02_重大差異_更新",
      ["重大差異(旧)", "旧:GitHub側", "更新状況", "本ブランチの対応/残り"],
      [
        ["GitHub側の同一性未確定", "比較対象が不明瞭", "解決", f"対象を {BRANCH} @ {HEAD} に確定。"],
        ["ws6/ws7(読取/編集)分離が違う", "1タブ内で表示/編集", "残存", "読取(ws6)/編集(ws7)の明示分離は未。表示ラベルで明確化予定。"],
        ["セル編集方式が違う", "ShiftPickerSheet選択", "解決", "ボトムシート選択で確定+触覚。誤巡回リスク解消。"],
        ["希望強制反映の有無", "なし", "残存", "『希望で上書き』(確認+Undo+ログ)は次段階で実装予定。"],
        ["実行中の進捗/中断表示", "反復/iter/s中心", "部分", "残り時間/満足度/未解決をライブ表示。背景最適化+完了通知で中断耐性を確保。中断秒数警告は未。"],
        ["自動保存の保存先/復元導線", "アプリ専用領域1.2s", "部分", "挙動は維持。保存先/復元範囲の画面文言明記は未。"],
        ["違反表示の重大度識別差", "赤枠/赤ドット", "部分", "『重大のみ』フィルタ追加。実線/破線の非色手がかりは未。"],
        ["チェック後の修正導線差", "状態/スコア中心", "解決", "コパイロットが原因family提示→編集タブへ誘導(NextAction相当)。"],
        ["Web soft breakdownの信頼性注意", "—", "解決", "ゴールデンテストでsoft完全一致はassertせず注記済(stale対策)。"],
        ["PC/スマホ表示の資料不足", "実機未確認", "部分", "トークン忠実モック(PNG/PDF)を作成し目視検証。実機スクショは別途。"],
      ],
      [26, 22, 12, 50], status_col=3)

# 03 片方のみの更新
sheet("03_片方のみ_更新",
      ["No", "対象", "旧:差異", "更新状況", "備考"],
      [
        [1, "月次入力ウィザード", "Webのみ", "残存", "希望→必要人数の月次フローは未。"],
        [2, "年次セットアップCTA", "Webのみ", "残存", "未。"],
        [3, "翌月切替バナー", "Webのみ", "残存", "未。"],
        [4, "希望サマリ/担当外警告", "Webのみ", "解決", "ホームに担当外希望警告+希望編集へ誘導を追加。"],
        [5, "NextActionBar(修正導線)", "Webのみ", "解決", "コパイロット→編集タブのジャンプを追加。"],
        [6, "中央RUN付きbottom-nav", "Webのみ", "残存", "意図的に不採用(下部コマンドバーで代替)。"],
        [7, "ws6/ws7分離", "Webのみ", "残存", "未。"],
        [8, "重大のみフィルタ", "Webのみ", "解決", "違反内訳に『重大のみ』トグルを追加。"],
        [9, "希望で上書き", "Webのみ", "残存", "次段階で実装予定(確認/Undo/ログ必須)。"],
        [10, "他の案(elitePool)", "Webのみ", "残存", "未。"],
        [11, "Wake Lock/中断警告", "Webのみ", "部分", "WorkManager背景最適化+完了通知で中断耐性は確保。Wake Lock固有/中断秒数警告は未。"],
        [12, "TapGame", "Webのみ", "残存", "意図的非採用(業務必須でない)。"],
        [13, "DefragLiveView(途中経過)", "Webのみ", "部分", "数値のライブ表示は実装。スケジュールのライブ描画は未。"],
        [14, "操作ログJSON書き出し", "Web広い", "解決", "追記式の操作監査ログ＋テキスト/JSON出力を実装。"],
        [15, "フッター変更履歴", "Webのみ", "残存", "未(開発者向けに分離想定)。"],
        [16, "テーマ(ライト/ダーク/高コントラスト)", "GitHubのみ", "解決(維持)", "Android固有として維持。"],
        [17, "片手モード", "GitHubのみ", "解決(維持)", "維持(下方寄せ)。"],
        [18, "Android Documents入出力", "GitHubのみ", "解決(維持)", "OSファイルピッカー維持。"],
        [19, "ネイティブHaptic", "GitHubのみ", "解決(追加)", "セル編集/選択で触覚を追加。"],
        [20, "AlertDialog色ピッカー", "GitHubのみ", "解決(維持)", "維持。"],
      ],
      [5, 24, 14, 12, 44], status_col=4)

# 04 残課題(優先)
sheet("04_残課題_優先",
      ["優先度", "残課題", "内容", "実装方針"],
      [
        ["済", "WorkManager+完了通知", "(実装済) 背景最適化+完了通知", "OptimizationWorker(CoroutineWorker)+Expedited+進捗Repository+通知。1ae0228。"],
        ["高", "希望で上書き", "希望シフトを勤務表へ強制反映", "確認ダイアログ(担当外時)+Undo+ログを必須化。VMにapplyWishes追加(着手中)。"],
        ["高", "プロセスkill完全耐性", "kill後もWorkerが入力を失わず再開", "stateをファイル受け渡しに(現状は同一プロセス参照→kill時失敗)。"],
        ["高", "ws6/ws7分離", "読取結果と編集中を明示分離", "勤務表タブに読取/編集ラベルとモード切替を追加、または統合時はラベル統一。"],
        ["中", "非色手がかり(実線/破線)", "HARD=実線/SOFT=破線で重大度識別", "違反セルの枠線スタイルをHARD/SOFTで分岐(色覚配慮)。"],
        ["中", "月次/年次ウィザード", "希望→必要人数の入力導線", "採否を判断のうえ編集タブにステップフローを追加。"],
        ["中", "他の案(elitePool)", "複数候補の比較選択", "エンジンが上位候補を保持→切替UI。"],
        ["低", "文言の完全統一", "状態語/ボタン名の統一案適用", "配布可/要確認/実行中/未計算などへ統一(churn配慮で個別に)。"],
        ["低", "フッター版数/変更履歴", "版数表示と履歴折りたたみ", "設定の最下部に版数、履歴は開発者向け。"],
      ],
      [8, 22, 34, 50])

# 05 根拠
sheet("05_根拠と前提(更新)",
      ["対象", "内容", "根拠/備考"],
      [
        ["比較対象(更新)", f"{BRANCH} @ {HEAD}", "本セッションでの一連コミット(c68fad5..e4f6de8)を反映。"],
        ["CI", "v6-engine-check (testDebugUnitTest + assembleDebug)", "各コミットで緑を確認(ゴールデンパリティ含む)。"],
        ["視覚検証", "tools/mock_render.py によるトークン忠実モック(PNG/PDF)", "実機エミュレータ非搭載のため、配色/角丸/タイポを忠実再現したモックで目視。"],
        ["Web golden", "soft breakdownはstale可能性。HARD/解決層のみを正準として比較。", "V6WebGoldenParityTest 参照。"],
        ["未確認", "実機スクリーンショット/操作動画", "実表示崩れ・モーダルアニメは別途実機確認が望ましい。"],
      ],
      [18, 50, 44])

path = "/home/user/MAGI-ShiftOptimizer/tools/magi_uiux_diff_report_updated.xlsx"
wb.save(path)
print("WROTE", path, "sheets:", wb.sheetnames)
